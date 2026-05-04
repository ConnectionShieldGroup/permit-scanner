#!/usr/bin/env python3
"""Processa XLSX do Reading MA — extrai permits qualificados.

Reading publica relatório mensal Excel via CivicPlus DocumentCenter.
Estrutura validada 03/mai (Feb 2026 file): 17 colunas, 127 linhas.

Aplica filtros padrão Reginaldo:
- Últimos 45 dias (cutoff)
- 7 categorias (skip outros)
- Skip closed/CO issued

Saída: JSON no formato Permit[] (CONTRACTS.md), pronto pro frontend.
"""
import json
import re
import sys
import uuid
from datetime import datetime, timezone, timedelta
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install openpyxl: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
INPUT_FILES = [
    ROOT.parent.parent / "../../tmp/reading-19481-February-2026.xlsx",
    ROOT.parent.parent / "../../tmp/reading-19305-January-2026.xlsx",
    ROOT.parent.parent / "../../tmp/reading-19061-December-2025.xlsx",
]
# Path correção — files estão em /tmp
INPUT_FILES = [
    Path("/tmp/reading-19481-February-2026.xlsx"),
    Path("/tmp/reading-19305-January-2026.xlsx"),
    Path("/tmp/reading-19061-December-2025.xlsx"),
]
OUTPUT = ROOT / "app" / "src" / "lib" / "reading-real.json"

# Heurística work_type (mesmas palavras-chave do Hingham)
def map_work_type(permit_for, description):
    text = f"{permit_for} {description}".lower()

    if any(k in text for k in ["new construction", "new home", "new single family",
                                "new dwelling", "construct new", "new const", "new house"]):
        return "new_construction"
    if any(k in text for k in ["kitchen remodel", "kitchen renovation", "kitchen reno", "kitchen"]):
        return "kitchen_renovation"
    if any(k in text for k in ["bathroom remodel", "bath renovation", "bath remodel",
                                "bathroom", "bath"]):
        return "bath_renovation"
    if any(k in text for k in ["room addition", "garage addition", "second story addition",
                                "addition"]):
        return "addition"
    if any(k in text for k in ["interior renovation", "residential renovation",
                                "commercial renovation", "basement remodel", "build-out",
                                "build out", "tenant fit-out", "tenant fit out",
                                "remodel", "alteration", "alter ", "renovation"]):
        return "renovation"
    if any(k in text for k in ["foundation permit", "foundation"]):
        return "foundation_permit"
    if any(k in text for k in ["building permit"]):
        return "building_permit"
    return None

def parse_date(v):
    """Converte valor de célula em ISO date."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, str):
        # Tenta YYYY-MM-DD HH:MM:SS
        m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", v.strip())
        if m:
            return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return None

def process_file(path, cutoff_date):
    if not path.exists():
        print(f"[skip] {path} não existe", file=sys.stderr)
        return []

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    out = []
    skipped_old = 0
    skipped_not_relevant = 0
    skipped_inactive = 0

    for row_num in range(2, ws.max_row + 1):
        # Lê todas células
        cells = [ws.cell(row=row_num, column=c).value for c in range(1, 18)]
        record_num, rec_type, title, st_no, st_name, app_name, app_email, app_phone, \
            owner, date_submitted, project_cost, total_paid, issued_date, doc_type, \
            permit_for, description, status = cells

        if not record_num:
            continue

        # Status: só Active (pula closed, withdrawn, etc)
        status_clean = (status or "").strip().lower()
        if status_clean and status_clean not in ("active", "issued", "open"):
            skipped_inactive += 1
            continue

        # Data: prefere issued_date, fallback date_submitted
        date_iso = parse_date(issued_date) or parse_date(date_submitted)
        if not date_iso or date_iso < cutoff_date:
            skipped_old += 1
            continue

        # Work type filter
        wt = map_work_type(permit_for or "", description or "")
        if wt is None:
            skipped_not_relevant += 1
            continue

        # Endereço
        addr_parts = []
        if st_no:
            addr_parts.append(str(st_no).strip())
        if st_name:
            addr_parts.append(str(st_name).strip().title())
        address = " ".join(addr_parts) if addr_parts else "Reading, MA"

        # Phone clean
        phone_clean = None
        if app_phone:
            phone_clean = re.sub(r"[^\d-]", "", str(app_phone)).strip("-")

        # Estimated value
        est_value = None
        if project_cost not in (None, "", 0, "0"):
            try:
                est_value = float(str(project_cost).replace(",", ""))
                if est_value <= 0:
                    est_value = None
            except (ValueError, TypeError):
                pass

        permit = {
            "id": str(uuid.uuid4()),
            "permit_number": str(record_num).strip(),
            "applicant_name": str(app_name).strip().title() if app_name else None,
            "address": address,
            "city": "Reading",
            "state": "MA",
            "phone": phone_clean,
            "email": str(app_email).strip().lower() if app_email else None,
            "work_type": wt,
            "permit_date": date_iso,
            "application_date": parse_date(date_submitted),
            "estimated_value": est_value,
            "status_source": status,
            "source_url": "https://www.readingma.gov/959/Monthly-Building-Permit-Report",
            "description": str(description).strip() if description else None,
            "raw_data": {
                "source": "reading-civicplus-xlsx",
                "permit_for": permit_for,
                "owner_name": str(owner).strip() if owner else None,
                "title": title,
                "doc_type": doc_type,
                "total_paid": total_paid,
                "month_file": path.name,
            },
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        out.append(permit)

    print(f"[{path.name}] {len(out)} qualified | "
          f"skipped: {skipped_old} old, {skipped_not_relevant} not relevant, "
          f"{skipped_inactive} inactive")
    return out

def main():
    # Reading publica com 2-3 meses de delay. Cutoff mais largo (120 dias)
    # captura último mês disponível. Pra produção, ajustar pra 90 dias.
    cutoff_dt = (datetime.now(timezone.utc) - timedelta(days=120)).date()
    cutoff_iso = cutoff_dt.isoformat()
    print(f"Cutoff: {cutoff_iso} (últimos 120 dias — Reading publica com delay)\n")

    all_permits = []
    for path in INPUT_FILES:
        permits = process_file(path, cutoff_iso)
        all_permits.extend(permits)

    # Dedup por permit_number (mantém o primeiro = mais recente, files ordenados)
    seen = set()
    deduped = []
    for p in all_permits:
        if p["permit_number"] not in seen:
            seen.add(p["permit_number"])
            deduped.append(p)

    # Sort por data desc
    deduped.sort(key=lambda p: p["permit_date"] or "1900-01-01", reverse=True)

    # Distribuição
    from collections import Counter
    by_type = Counter(p["work_type"] for p in deduped)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT, "w") as f:
        json.dump(deduped, f, indent=2, ensure_ascii=False, default=str)

    print(f"\n[ok] {len(deduped)} unique Reading permits saved to {OUTPUT}")
    print(f"by_work_type: {dict(by_type)}")

    if deduped:
        sample = deduped[0]
        print(f"\nSample (most recent):")
        print(f"  {sample['permit_number']} - {sample['address']}")
        print(f"  {sample['applicant_name']} ({sample['phone']} / {sample['email']})")
        print(f"  ${sample['estimated_value']:,.0f}" if sample['estimated_value'] else "  no value")
        print(f"  {sample['description'][:120]}...")

if __name__ == "__main__":
    main()
